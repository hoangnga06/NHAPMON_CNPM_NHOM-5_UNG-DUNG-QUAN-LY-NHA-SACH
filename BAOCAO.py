import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

BOOK_FILE = "books.json"
SALE_FILE = "sales.json"

# ======================
# LOAD DATA
# ======================
def load_books():
    if not os.path.exists(BOOK_FILE):
        return {}
    with open(BOOK_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def load_sales():
    if not os.path.exists(SALE_FILE):
        return []
    with open(SALE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

# =================================================
# BÁO CÁO TỒN KHO
# =================================================
def report_inventory():
    books = load_books()
    if not books:
        print("❌ Chưa có dữ liệu sách")
        return

    # Gom sách theo thể loại
    categories = {}
    for bid, b in books.items():
        cat = b.get("category", "KHÁC").upper()
        categories.setdefault(cat, [])
        categories[cat].append((bid, b))

    # Sắp xếp thể loại theo ABC
    for cat in sorted(categories.keys()):
        print("\n" + "=" * 100)
        print(f"📚 THỂ LOẠI: {cat}")
        print("=" * 100)
        print("{:<8} {:<30} {:<10} {}".format(
            "Mã", "Tên sách", "Tồn", "Cảnh báo"
        ))
        print("-" * 100)

        for bid, b in categories[cat]:
            qty = b.get("qty", 0)
            warn = "⚠️ Tồn thấp" if qty < 5 else ""
            print("{:<8} {:<30} {:<10} {}".format(
                bid,
                b.get("name", ""),
                qty,
                warn
            ))
# =================================================
# BÁO CÁO DOANH THU
# =================================================
def report_revenue():
    sales = load_sales()
    if not sales:
        print("❌ Chưa có dữ liệu bán hàng")
        return

    stats = {}

    for s in sales:
        date = s["time"].split(" ")[0]   # dd/mm/yyyy
        stats.setdefault(date, 0)
        stats[date] += s.get("pay", 0)

    print("\n💰 BÁO CÁO DOANH THU THEO NGÀY")
    print("-" * 40)
    for d in sorted(stats):
        print(f"{d}: {stats[d]:,.0f} đ")
def report_revenue_by_month():
    sales = load_sales()
    if not sales:
        print("❌ Chưa có dữ liệu bán hàng")
        return

    stats = {}

    for s in sales:
        # time: dd/mm/yyyy hh:mm
        d = datetime.strptime(s["time"], "%d/%m/%Y %H:%M")
        key = d.strftime("%m/%Y")  # tháng/năm
        stats.setdefault(key, 0)
        stats[key] += s.get("pay", 0)

    print("\n💰 BÁO CÁO DOANH THU THEO THÁNG")
    print("-" * 40)
    for k in sorted(stats):
        print(f"{k}: {stats[k]:,.0f} đ")
def report_revenue_by_year():
    sales = load_sales()
    if not sales:
        print("❌ Chưa có dữ liệu bán hàng")
        return

    stats = {}

    for s in sales:
        d = datetime.strptime(s["time"], "%d/%m/%Y %H:%M")
        year = d.year
        stats.setdefault(year, 0)
        stats[year] += s.get("pay", 0)

    print("\n💰 BÁO CÁO DOANH THU THEO NĂM")
    print("-" * 40)
    for y in sorted(stats):
        print(f"{y}: {stats[y]:,.0f} đ")

# =================================================
# XUẤT EXCEL
# =================================================

EXCEL_FILE = "BAOCAO_TONKHO.xlsx"

def update_inventory_excel():
    books = load_books()
    if not books:
        print("❌ Không có dữ liệu sách")
        return

    # Nếu chưa có file → tạo mới
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "TonKho"
        ws.append(["Mã sách", "Tên sách", "Thể loại", "Tồn kho", "Cảnh báo"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["TonKho"]

    # Xoá dữ liệu cũ (giữ lại header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # Ghi dữ liệu mới
    for bid, b in books.items():
        qty = b.get("qty", 0)
        warn = "⚠ Tồn thấp" if qty < 5 else ""
        ws.append([
            bid,
            b.get("name", ""),
            b.get("category", ""),
            qty,
            warn
        ])

    wb.save(EXCEL_FILE)
    print(f"✅ Đã cập nhật vào file {EXCEL_FILE}")
from openpyxl import Workbook, load_workbook

def export_all_report_excel_one_sheet():
    books = load_books()
    sales = load_sales()

    if not books and not sales:
        print("❌ Không có dữ liệu để xuất")
        return

    fname = "BAOCAO_TONG_HOP_1SHEET.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BAO_CAO"

    row = 1
    # =============================
    # BÁO CÁO TỒN KHO
    # =============================
    ws.cell(row=row, column=1, value="BÁO CÁO TỒN KHO")
    row += 1

    ws.append(["Mã sách", "Tên sách", "Thể loại", "Tồn kho", "Cảnh báo"])
    row += 1

    for bid, b in books.items():
        qty = b.get("qty", 0)
        warn = "Tồn thấp" if qty < 5 else ""
        ws.append([bid, b.get("name",""), b.get("category",""), qty, warn])
        row += 1

    row += 2  # dòng trống

    # =============================
    # DOANH THU THEO NGÀY
    # =============================
    ws.cell(row=row, column=1, value="DOANH THU THEO NGÀY")
    row += 1
    ws.append(["Ngày", "Doanh thu"])
    row += 1

    stats_day = {}
    for s in sales:
        day = s["time"].split(" ")[0]
        stats_day.setdefault(day, 0)
        stats_day[day] += s.get("pay", 0)

    for d in sorted(stats_day):
        ws.append([d, stats_day[d]])
        row += 1

    row += 2

    # =============================
    # DOANH THU THEO THÁNG
    # =============================
    ws.cell(row=row, column=1, value="DOANH THU THEO THÁNG")
    row += 1
    ws.append(["Tháng", "Doanh thu"])
    row += 1

    stats_month = {}
    for s in sales:
        d = datetime.strptime(s["time"], "%d/%m/%Y %H:%M")
        key = d.strftime("%m/%Y")
        stats_month.setdefault(key, 0)
        stats_month[key] += s.get("pay", 0)

    for m in sorted(stats_month):
        ws.append([m, stats_month[m]])
        row += 1

    row += 2

    # =============================
    # DOANH THU THEO NĂM
    # =============================
    ws.cell(row=row, column=1, value="DOANH THU THEO NĂM")
    row += 1
    ws.append(["Năm", "Doanh thu"])
    row += 1

    stats_year = {}
    for s in sales:
        y = datetime.strptime(s["time"], "%d/%m/%Y %H:%M").year
        stats_year.setdefault(y, 0)
        stats_year[y] += s.get("pay", 0)

    for y in sorted(stats_year):
        ws.append([y, stats_year[y]])
        row += 1

    wb.save(fname)
    print(f"✅ Đã xuất báo cáo tổng hợp 1 sheet: {fname}")

def report_customer_purchase_detail():
    sales = load_sales()
    if not sales:
        print("❌ Chưa có dữ liệu bán hàng")
        return

    phone = input("Nhập SĐT khách hàng: ").strip()

    found = False
    print("\n🧾 CHI TIẾT KHÁCH HÀNG ĐÃ MUA")
    print("=" * 80)

    for s in sales:
        cus = s.get("customer", {})
        if cus.get("phone") == phone:
            found = True
            print(f"Mã đơn   : {s.get('id')}")
            print(f"Thời gian: {s.get('time')}")
            print(f"Khách    : {cus.get('name')}")
            print("-" * 80)
            print("{:<25} {:>5} {:>12}".format("Tên sách", "SL", "Thành tiền"))

            for item in s.get("items", []):
                total_item = item["price"] * item["qty"]
                print("{:<25} {:>5} {:>12,.0f}".format(
                    item["name"],
                    item["qty"],
                    total_item
                ))

            print("-" * 80)
            print(f"Tổng tiền : {s.get('total', 0):,.0f}")
            print(f"Giảm giá  : {s.get('discount', 0):,.0f}")
            print(f"Khách trả : {s.get('pay', 0):,.0f}")
            print("=" * 80)

    if not found:
        print("❌ Không tìm thấy giao dịch của khách hàng này")
# =================================================
# MENU BÁO CÁO (ADMIN)
# =================================================
def menu_baocao():
    while True:
        print("\n=== 📊 MENU BÁO CÁO ===")
        print("1. Xem báo cáo tồn kho (console)")
        print("2. Xem doanh thu (console)")
        print("3. Xuất BÁO CÁO TỔNG HỢP (Excel)")
        print("0. Quay lại")

        ch = input("Chọn: ")

        if ch == "1":
            report_inventory()

        elif ch == "2":
            print("\n--- DOANH THU ---")
            report_revenue()
            report_revenue_by_month()
            report_revenue_by_year()

        elif ch == "3":
            export_all_report_excel_one_sheet()# ✅ FILE TỔNG HỢP DUY NHẤT

        elif ch == "0":
            break
        else:
            print("❌ Lựa chọn không hợp lệ")
